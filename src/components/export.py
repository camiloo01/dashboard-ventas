"""
export.py — Exportación de datos en Excel con múltiples hojas y totales.
"""

from __future__ import annotations

import io
import pandas as pd
import streamlit as st

from src.utils.helpers import pct_change, sign


def render_export(fa: pd.DataFrame, fb: pd.DataFrame, table_df: pd.DataFrame, date_filter: dict | None = None) -> None:
    if table_df.empty and fa.empty and fb.empty:
        return
    st.markdown('<div class="section-title">⬇️ Exportar</div>', unsafe_allow_html=True)
    excel_data = _build_excel(fa, fb, table_df, date_filter)

    # Nombre del archivo con fechas si hay filtro de día
    if date_filter:
        nombre = f"reporte_{date_filter['fecha_a']}_vs_{date_filter['fecha_b']}.xlsx"
    else:
        nombre = "comparativa_ventas.xlsx"

    st.download_button(
        label="📥 Descargar reporte completo en Excel",
        data=excel_data,
        file_name=nombre,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _build_excel(fa: pd.DataFrame, fb: pd.DataFrame, table_df: pd.DataFrame, date_filter: dict | None = None) -> bytes:
    buffer = io.BytesIO()

    sheets = {
        "Fabricante":  "Fabricante",
        "Region":      "Region",
        "Tienda":      "Descripción Centro",
        "Canal":       "Canal",
    }

    hojas = {}

    # ── Hoja Resumen siempre presente ─────────────────────────────────────────
    resumen_rows = []

    if date_filter:
        resumen_rows += [
            {"Descripción": "Fecha Año Anterior", "Valor": str(date_filter["fecha_a"])},
            {"Descripción": "Fecha Año Actual",   "Valor": str(date_filter["fecha_b"])},
            {"Descripción": "", "Valor": ""},
        ]

    resumen_rows += [
        {"Descripción": "Total Filas Año Anterior",     "Valor": f"{len(fa):,}"},
        {"Descripción": "Total Filas Año Actual",       "Valor": f"{len(fb):,}"},
        {"Descripción": "",                             "Valor": ""},
        {"Descripción": "Ingresos Año Anterior",        "Valor": f"$ {fa['_ingreso'].sum():,.0f}"},
        {"Descripción": "Ingresos Año Actual",          "Valor": f"$ {fb['_ingreso'].sum():,.0f}"},
        {"Descripción": "Variación $",                  "Valor": f"{sign(fb['_ingreso'].sum() - fa['_ingreso'].sum())}$ {fb['_ingreso'].sum() - fa['_ingreso'].sum():,.0f}"},
        {"Descripción": "Variación %",                  "Valor": f"{sign(pct_change(fa['_ingreso'].sum(), fb['_ingreso'].sum()))}{pct_change(fa['_ingreso'].sum(), fb['_ingreso'].sum())}%"},
    ]

    # Unidades si existe la columna
    cant_col = _find_col(fa, ["Cantidad", "cantidad", "Qty", "Units"])
    if cant_col:
        resumen_rows += [
            {"Descripción": "",                          "Valor": ""},
            {"Descripción": "Unidades Año Anterior",     "Valor": f"{int(fa[cant_col].sum()):,}"},
            {"Descripción": "Unidades Año Actual",       "Valor": f"{int(fb[cant_col].sum()):,}"},
        ]

    hojas["Resumen"] = pd.DataFrame(resumen_rows)

    # ── Resto de hojas ────────────────────────────────────────────────────────
    if not table_df.empty:
        hojas["Fabricante"] = table_df

    for sheet_name, col in sheets.items():
        if sheet_name == "Fabricante":
            continue
        if col not in fa.columns:
            continue
        df = _build_summary(fa, fb, col)
        if not df.empty:
            hojas[sheet_name] = df

    if len(hojas) == 1:  # solo Resumen
        hojas["Sin datos"] = pd.DataFrame({"Mensaje": ["No hay datos con los filtros actuales"]})

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in hojas.items():
            _write_sheet(writer, df, sheet_name, add_total=sheet_name != "Resumen")

    return buffer.getvalue()


def _build_summary(fa: pd.DataFrame, fb: pd.DataFrame, group_col: str) -> pd.DataFrame:
    cant_col = _find_col(fa, ["Cantidad", "cantidad", "Qty", "Units"])

    agg_dict = {"ingreso": ("_ingreso", "sum")}
    if cant_col:
        agg_dict["cant"] = (cant_col, "sum")

    agg_a = fa.groupby(group_col).agg(**agg_dict)
    agg_b = fb.groupby(group_col).agg(**agg_dict)
    items = sorted(set(agg_a.index) | set(agg_b.index))

    if not items:
        return pd.DataFrame()

    rows = []
    for item in items:
        ia = agg_a.loc[item, "ingreso"] if item in agg_a.index else 0
        ib = agg_b.loc[item, "ingreso"] if item in agg_b.index else 0
        ca = agg_a.loc[item, "cant"]    if (cant_col and item in agg_a.index) else "-"
        cb = agg_b.loc[item, "cant"]    if (cant_col and item in agg_b.index) else "-"
        dp = pct_change(ia, ib)

        rows.append({
            group_col:       item,
            "Ing. Año Ant.": f"$ {ia:,.0f}",
            "Ing. Año Act.": f"$ {ib:,.0f}",
            "Variacion $":   f"{sign(ib - ia)}$ {ib - ia:,.0f}",
            "Variacion %":   f"{sign(dp)}{dp}%",
            "Unid. Ant.":    str(int(ca)) if ca != "-" else "-",
            "Unid. Act.":    str(int(cb)) if cb != "-" else "-",
        })

    return pd.DataFrame(rows).sort_values("Ing. Año Act.", ascending=False).reset_index(drop=True)


def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, add_total: bool = False) -> None:
    """Escribe un DataFrame en una hoja con columnas ajustadas y fila de totales."""

    # Agregar fila de TOTAL al final
    if add_total and not df.empty:
        total_row = {col: "" for col in df.columns}
        first_col = df.columns[0]
        total_row[first_col] = "TOTAL"

        # Sumar columnas de ingreso numéricamente
        for col in df.columns:
            if "Ing." in col or "Unid." in col:
                try:
                    valores = df[col].str.replace(r"[\$\s,+]", "", regex=True).str.replace(r"[^\d\.\-]", "", regex=True)
                    total = pd.to_numeric(valores, errors="coerce").sum()
                    if "Ing." in col:
                        total_row[col] = f"$ {total:,.0f}"
                    else:
                        total_row[col] = str(int(total))
                except Exception:
                    pass

        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

    df.to_excel(writer, index=False, sheet_name=sheet_name)

    # Ajustar ancho de columnas
    worksheet = writer.sheets[sheet_name]
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(
            df[col].astype(str).map(len).max(),
            len(str(col))
        ) + 4
        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
        worksheet.column_dimensions[col_letter].width = min(max_len, 40)

    # Poner la fila de TOTAL en negrita
    if add_total:
        from openpyxl.styles import Font, PatternFill
        last_row = worksheet.max_row
        for cell in worksheet[last_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2a2a3d", end_color="2a2a3d", fill_type="solid")


def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    return None